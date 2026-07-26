from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
import unicodedata
from zipfile import BadZipFile, ZipFile


QUESTION_HEADER = "Câu hỏi của người dùng"
EXPECTED_HEADER = "Kết quả mong đợi"
MAX_WORKBOOK_BYTES = 5 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 25 * 1024 * 1024
MAX_EVALUATION_CASES = 500


class WorkbookValidationError(ValueError):
    """The uploaded workbook does not match the evaluation contract."""


@dataclass(frozen=True, slots=True)
class EvaluationCaseInput:
    row_number: int
    question: str
    expected: str


@dataclass(frozen=True, slots=True)
class ParsedEvaluationWorkbook:
    sheet_name: str
    cases: tuple[EvaluationCaseInput, ...]


def parse_evaluation_workbook(
    filename: str,
    content: bytes,
) -> ParsedEvaluationWorkbook:
    if Path(filename).suffix.lower() != ".xlsx":
        raise WorkbookValidationError("Chỉ hỗ trợ file Excel định dạng .xlsx.")
    if not content:
        raise WorkbookValidationError("File Excel đang trống.")
    if len(content) > MAX_WORKBOOK_BYTES:
        raise WorkbookValidationError("File Excel vượt quá giới hạn 5 MB.")
    _validate_zip_size(content)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel evaluation requires openpyxl. Install Backend dependencies."
        ) from exc

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise WorkbookValidationError("Không thể đọc file Excel đã tải lên.") from exc

    try:
        worksheet = _matching_worksheet(workbook.worksheets)
        cases: list[EvaluationCaseInput] = []
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, max_col=2, values_only=True),
            start=2,
        ):
            question = _cell_text(values[0])
            expected = _cell_text(values[1])
            if not question and not expected:
                continue
            if not question or not expected:
                missing = QUESTION_HEADER if not question else EXPECTED_HEADER
                raise WorkbookValidationError(
                    f"Dòng {row_number} đang thiếu dữ liệu cột “{missing}”."
                )
            if len(question) > 4000 or len(expected) > 4000:
                raise WorkbookValidationError(
                    f"Dòng {row_number} vượt quá giới hạn 4.000 ký tự mỗi ô."
                )
            cases.append(
                EvaluationCaseInput(
                    row_number=row_number,
                    question=question,
                    expected=expected,
                )
            )
            if len(cases) > MAX_EVALUATION_CASES:
                raise WorkbookValidationError(
                    f"Mỗi lần chỉ đánh giá tối đa {MAX_EVALUATION_CASES} test case."
                )
        if not cases:
            raise WorkbookValidationError("File Excel không có test case để đánh giá.")
        return ParsedEvaluationWorkbook(
            sheet_name=worksheet.title,
            cases=tuple(cases),
        )
    finally:
        workbook.close()


def _matching_worksheet(worksheets):
    expected = (_normalize_header(QUESTION_HEADER), _normalize_header(EXPECTED_HEADER))
    for worksheet in worksheets:
        values = next(
            worksheet.iter_rows(min_row=1, max_row=1, max_col=3, values_only=True),
            (),
        )
        actual = tuple(_normalize_header(_cell_text(value)) for value in values[:2])
        has_extra_header = len(values) > 2 and bool(_cell_text(values[2]))
        if actual == expected and not has_extra_header:
            return worksheet
    raise WorkbookValidationError(
        "Không tìm thấy sheet có đúng hai cột “Câu hỏi của người dùng” "
        "và “Kết quả mong đợi”."
    )


def _validate_zip_size(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            total_size = sum(item.file_size for item in archive.infolist())
    except BadZipFile as exc:
        raise WorkbookValidationError("File tải lên không phải workbook .xlsx hợp lệ.") from exc
    if total_size > MAX_UNCOMPRESSED_BYTES:
        raise WorkbookValidationError(
            "Dữ liệu giải nén của file Excel vượt quá giới hạn 25 MB."
        )


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return re.sub(r"\s+", " ", normalized).strip()
